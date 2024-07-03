import openpyxl
import pandas as pd

from sqlalchemy import create_engine

# 打开Excel文件
workbook = openpyxl.load_workbook(
    "C:/Users/WB-wangyu/Desktop/引流导管BOM小程序数据库.xlsx"
)

# 获取所有sheet的名称
sheet_names = workbook.sheetnames


# DC数据处理：选择第三个sheet
sheet_dc = workbook[sheet_names[2]]

# 获取所有行的数据
rows_dc = sheet_dc.iter_rows(values_only=True)

# 获取第一行作为列名
columns_dc = [cell for cell in next(rows_dc)]

# 将剩余行的数据保存到列表中
data_dc = [row for row in rows_dc]

# 创建DataFrame对象
table_yinliu = pd.DataFrame(data_dc, columns=columns_dc)

# # 删除第一列
# table_yinliu = table_yinliu.drop(table_yinliu.columns[0], axis=1)

columns_dc = [
    "id",
    "MaterialCode",
    "fileCode",
    "Name",
    "ProductSpecifications",
    "Version",
]

table_yinliu.columns = columns_dc


for i in range(len(table_yinliu)):
    Product = table_yinliu.loc[i, "ProductSpecifications"]
    # 将 "/" 替换为 "-"
    Product = Product.replace("/", "-")
    # 按照 "-" 拆分字符串
    split_string = Product.split("-")

    # 设置产品代号
    table_yinliu.loc[i, "ProductCode"] = split_string[0]

    # 设置直径、长度
    yinliuDiameter = split_string[1][0:2]
    yinliuLength = split_string[1][2:4]
    table_yinliu.loc[i, "yinliuDiameter"] = yinliuDiameter
    table_yinliu.loc[i, "yinliuLength"] = yinliuLength

    # 设置锁定形式
    yinliuLockStyle = split_string[2][0:1]
    table_yinliu.loc[i, "yinliuLockStyle"] = yinliuLockStyle

    # 设置头端样式
    yinliuHeadStyle = split_string[2][1:]
    table_yinliu.loc[i, "yinliuHeadStyle"] = yinliuHeadStyle

    # 设置配置代号
    table_yinliu.loc[i, "yinliuConfigurationCode"] = split_string[3]

table_yinliu = table_yinliu.astype(str)


# 创建 SQLAlchemy 引擎
engine = create_engine("mysql+pymysql://root:123456@localhost:3306/yinliu")

# 将id列转换为整数类型
table_yinliu["id"] = table_yinliu["id"].astype(int)

table_yinliu.to_sql("dc_yinliu", con=engine, if_exists="replace", index=False)


# BDC数据处理： 选择第二个sheet
sheet_bdc = workbook[sheet_names[1]]

# 获取所有行的数据
rows_bdc = sheet_bdc.iter_rows(values_only=True)

# 获取第一行作为列名
columns_bdc = [cell for cell in next(rows_bdc)]

# 将剩余行的数据保存到列表中
data_bdc = [row for row in rows_bdc]

# 创建DataFrame对象
table_dandao = pd.DataFrame(data_bdc, columns=columns_bdc)

# # 删除第一列
# table_dandao = table_dandao.drop(table_dandao.columns[0], axis=1)

columns_bdc = [
    "id",
    "MaterialCode",
    "fileCode",
    "Name",
    "ProductSpecifications",
    "Version",
]

table_dandao.columns = columns_bdc

for i in range(len(table_dandao)):
    Product = table_dandao.loc[i, "ProductSpecifications"]
    # 将 "/" 替换为 "-"
    Product = Product.replace("/", "-")
    # 按照 "-" 拆分字符串
    split_string = Product.split("-")

    # 设置产品代号
    table_dandao.loc[i, "ProductCode"] = split_string[0]

    # 设置直径、长度、内外引流方式
    if len(split_string[1]) == 4:
        dandaoDiameter = split_string[1][0:2]
        dandaoLength = split_string[1][2:4]
        table_dandao.loc[i, "dandaoDiameter"] = dandaoDiameter
        table_dandao.loc[i, "dandaoLength"] = dandaoLength
        table_dandao.loc[i, "drainageMethod"] = None
    else:
        dandaoDiameter = split_string[1][0:2]
        dandaoLength = split_string[1][2:4]
        drainageMethod = split_string[1][4:]
        table_dandao.loc[i, "dandaoDiameter"] = dandaoDiameter
        table_dandao.loc[i, "dandaoLength"] = dandaoLength
        table_dandao.loc[i, "drainageMethod"] = drainageMethod

    # 设置头端样式
    table_dandao.loc[i, "dandaoHeadStyle"] = split_string[2]

    # 设置配置代号
    table_dandao.loc[i, "dandaoConfigurationCode"] = split_string[3]

table_dandao = table_dandao.astype(str)


# 创建 SQLAlchemy 引擎
engine = create_engine("mysql+pymysql://root:123456@localhost:3306/yinliu")

# 将id列转换为整数类型
table_dandao["id"] = table_dandao["id"].astype(int)

table_dandao.to_sql("dc_dandao", con=engine, if_exists="replace", index=False)


def add_leading_zeros_to_string(input_string):
    numbers = input_string.split(",")
    formatted_numbers = [f"{int(num):02d}" for num in numbers]
    return ",".join(formatted_numbers)


# 获取第一个sheet
sheet_component = workbook.worksheets[0]

# 获取所有合并单元格
merged_cells_ranges = sheet_component.merged_cells.ranges

# 创建DataFrame对象
table_component = pd.DataFrame(sheet_component.values)

# 将每个合并的单元格拆分开
for merged_range in merged_cells_ranges:
    min_row, min_col, max_row, max_col = (
        merged_range.min_row,
        merged_range.min_col,
        merged_range.max_row,
        merged_range.max_col,
    )
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if row != min_row or col != min_col:
                table_component.iloc[row - 1, col - 1] = table_component.iloc[
                    min_row - 1, min_col - 1
                ]

# 关闭Excel文件
workbook.close()

# 创建DataFrame对象
# table_component = pd.DataFrame(data, columns=columns)

# 选择从第三行开始的数据
table_component = table_component.iloc[2:, :]

# 重置索引
table_component.reset_index(drop=True, inplace=True)

for i in range(len(table_component)):
    for j in range(13):
        if table_component.iloc[i, j] == None:
            table_component.iloc[i, j] = ""

# 将胆道直径修改为标准格式
dandaoDiameter = table_component.iloc[:, 2].astype(str)
for i in range(len(dandaoDiameter)):
    if dandaoDiameter[i].strip() != "":
        dandaoDiameter[i] = add_leading_zeros_to_string(dandaoDiameter[i])
table_component.iloc[:, 2] = dandaoDiameter

# 将引流直径修改为标准格式
yinliuDiameter = table_component.iloc[:, 7].astype(str)
for i in range(len(yinliuDiameter)):
    if yinliuDiameter[i].strip() != "":
        yinliuDiameter[i] = add_leading_zeros_to_string(yinliuDiameter[i])
table_component.iloc[:, 7] = yinliuDiameter

table_component.insert(0, "id", range(1, len(table_component) + 1))

# 重置列名
columns_component = [
    "id",
    "ProductCode",
    "dandaoDiameter",
    "dandaoLength",
    "drainageMethod",
    "dandaoHeadStyle",
    "dandaoConfigurationCode",
    "yinliuDiameter",
    "yinliuLength",
    "yinliuLockStyle",
    "yinliuHeadStyle",
    "yinliuConfigurationCode",
    "subset",
    "majorCategory",
    "materialCode",
    "drawingCode",
    "Name",
    "specification",
    "material",
    "color",
    "numbers",
    "unit",
    "materialCategory",
    "Note",
    "perPrice",
    "totalPrice",
]

table_component.columns = columns_component

# 设置总价为 单价 * 数量
table_component["totalPrice"] = table_component["perPrice"].astype(float) * table_component["numbers"].astype(float)
table_component["totalPrice"] = table_component["totalPrice"].round(4)

# 创建 SQLAlchemy 引擎
engine = create_engine("mysql+pymysql://root:123456@localhost:3306/yinliu")

# 将id列转换为整数类型
table_component["id"] = table_component["id"].astype(int)

table_component.to_sql("components", con=engine, if_exists="replace", index=False)
