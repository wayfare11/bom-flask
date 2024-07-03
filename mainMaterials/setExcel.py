import openpyxl
import unicodedata
from openpyxl.styles import Alignment

def column_length(sheet):

    sheet_name = sheet.title

    start_row = 3
    addition_length = 2
    if sheet_name == "附录":
        start_row = 1
        addition_length = 5
    
    # 遍历每一列，根据每列内各个字符串的最大长度调整列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column[start_row:]:
            try:  # 部分单元格可能为空，需要进行异常处理
                cell_value = str(cell.value)
                cell_length = sum(2 if unicodedata.east_asian_width(c) in 'FWA' else 1 for c in cell_value)  # 根据字符类型计算字符串长度
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = max_length + addition_length
        sheet.column_dimensions[column_letter].width = adjusted_width

def setExcel(sheet, index_packaging, index_auxiliary):

    # 设置数据居中显示
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置第一行的字体大小为18
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(size=18)

    # 设置第四行的字体加粗
    for cell in sheet[4]:
        cell.font = openpyxl.styles.Font(bold=True)

    # 设置2、3、4行的行高与第一行一致
    for row in range(2, 5):
        sheet.row_dimensions[row].height = 23

    for cell in sheet[f'A{index_packaging}:K{index_packaging}'][0]:
        cell.font = openpyxl.styles.Font(bold=True)

    for cell in sheet[f'A{index_auxiliary}:K{index_auxiliary}'][0]:
        cell.font = openpyxl.styles.Font(bold=True)

    # 设置合并单元格和对齐方式
    sheet.merge_cells('A1:K1')
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('A2:E2')
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')

    sheet.merge_cells('F2:K2')
    sheet['F2'].alignment = Alignment(horizontal='left', vertical='center')

    sheet.merge_cells('A3:E3')
    sheet['A3'].alignment = Alignment(horizontal='left', vertical='center')

    sheet.merge_cells('F3:K3')
    sheet['F3'].alignment = Alignment(horizontal='left', vertical='center')

    sheet.merge_cells(f'A{index_packaging}:K{index_packaging}')
    sheet[f'A{index_packaging}'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells(f'A{index_auxiliary}:K{index_auxiliary}')
    sheet[f'A{index_auxiliary}'].alignment = Alignment(horizontal='center', vertical='center')

    # 为单元格添加边框
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

    column_length(sheet)

def setAppendixExcel(sheet):
    # 设置数据居中显示
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置第一行的字体大小为14，并加粗，其余行的字体大小为12
    # 设置第一行的字体大小为18
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(size=14, bold = True)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = openpyxl.styles.Font(size=12)

    # 合并单元格并设置合并后的单元格对齐方式
    sheet.merge_cells('A1:D1')
    sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')

    # 为单元格添加边框
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

    column_length(sheet)