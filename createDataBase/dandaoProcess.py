import openpyxl
import pandas as pd

from sqlalchemy import create_engine, inspect
from sqlalchemy.sql import text

# 打开Excel文件
workbook = openpyxl.load_workbook('C:/Users/WB-wangyu/Desktop/文档/开发文档/引流导管BOM工具/引流导管BOM小程序数据库.xlsx')

# 获取所有sheet的名称
sheet_names = workbook.sheetnames

# 选择第二个sheet
sheet = workbook[sheet_names[1]]

# 获取所有行的数据
rows = sheet.iter_rows(values_only=True)

# 获取第一行作为列名
columns = [cell for cell in next(rows)]

# 将剩余行的数据保存到列表中
data = [row for row in rows]

# 关闭Excel文件
workbook.close()

# 创建DataFrame对象
table_dandao = pd.DataFrame(data, columns=columns)

# # 删除第一列
# table_dandao = table_dandao.drop(table_dandao.columns[0], axis=1)

columns = ['id', 'MaterialCode', 'fileCode', 'Name', 'ProductSpecifications', 'Version']

table_dandao.columns = columns

for i in range(len(table_dandao)):
    Product = table_dandao.loc[i,"ProductSpecifications"]
    # 将 "/" 替换为 "-"
    Product = Product.replace("/", "-")
    # 按照 "-" 拆分字符串
    split_string = Product.split("-")

    # 设置产品代号
    table_dandao.loc[i,"ProductCode"] = split_string[0]

    # 设置直径、长度、内外引流方式
    if len(split_string[1]) == 4:
        dandaoDiameter = split_string[1][0:2]
        dandaoLength = split_string[1][2:4]
        table_dandao.loc[i,"dandaoDiameter"] = dandaoDiameter
        table_dandao.loc[i,"dandaoLength"] = dandaoLength
        table_dandao.loc[i,"drainageMethod"] = None
    else:
        dandaoDiameter = split_string[1][0:2]
        dandaoLength = split_string[1][2:4]
        drainageMethod = split_string[1][4:]
        table_dandao.loc[i,"dandaoDiameter"] = dandaoDiameter
        table_dandao.loc[i,"dandaoLength"] = dandaoLength
        table_dandao.loc[i,"drainageMethod"] = drainageMethod

    # 设置头端样式
    table_dandao.loc[i,"dandaoHeadStyle"] = split_string[2]

    # 设置配置代号
    table_dandao.loc[i,"dandaoConfigurationCode"] = split_string[3]

table_dandao = table_dandao.astype(str)



# 创建 SQLAlchemy 引擎
engine = create_engine('mysql+pymysql://root:123456@10.16.0.15:3306/yinliu')

# 将id列转换为整数类型
table_dandao['id'] = table_dandao['id'].astype(int)

table_dandao.to_sql('dandao', con=engine, if_exists='replace', index=False)


# 创建 SQLAlchemy 引擎
engine = create_engine('mysql+pymysql://root:123456@localhost:3306/yinliu')

# 将id列转换为整数类型
table_dandao['id'] = table_dandao['id'].astype(int)

table_dandao.to_sql('dandao', con=engine, if_exists='replace', index=False)